# Name: import_schemas_and_domains.py
# Author: Conor MacNaughton - conor.macnaughton@gov.bc.ca
# Created: 2021-06-15
# Last Updated: 2024-04-22
# Purpose: This tool takes input schema and domain Excel files, producing a point feature class or 
#          domain table for each sheet within them. The domains are then assigned to the appropriate 
#          fields within the feature classes.
# Inputs: 1) path to existing or new file geodatabase
#         2) path to Excel holding feature class schemas
#         3) path to Excel holding domains
# Outputs: feature classes and domains in geodatabase, with domains assigned to appropriate fields


import os
from openpyxl import load_workbook
import arcpy as ap

# Paths
out_gdb = r''
schema_excel = r''
domains_excel = r''

out_folder_path = os.path.dirname(out_gdb)
out_gdb_name = os.path.basename(out_gdb)

try:
    # Create geodatabase if it doesn't exist
    if not os.path.exists(out_gdb):
        ap.management.CreateFileGDB(out_folder_path, out_gdb_name)
        print(f'\nGeodatabase created: {out_gdb}')

    # Set workspace
    ap.env.workspace = out_gdb
    ap.env.overwriteOutput = True

    # Function to import tables from Excel
    def import_tables_from_excel(excel_file):
        workbook = load_workbook(filename=excel_file)
        sheets = workbook.sheetnames
        print(f'{len(sheets)} found: {sheets}')
        for sheet in sheets:
            out_table = os.path.join(out_gdb, sheet)
            ap.conversion.ExcelToTable(excel_file, out_table, sheet)

    # Import schema tables
    print('\nImporting schema tables...')
    import_tables_from_excel(schema_excel)

    # Create feature classes from tables
    for table in ap.ListTables():
        out_fc = table + "_FC"
        ap.management.CreateFeatureclass(out_gdb, out_fc, "POINT")
        ap.management.JoinField(out_fc, "OBJECTID", table, "OBJECTID")
        ap.management.DeleteRows(out_fc)
        ap.management.Delete(table)
        print(f'{out_fc} feature class created')

    # Import domain tables and assign domains
    print('\nImporting domain tables...')
    import_tables_from_excel(domains_excel)
    for sheet in ap.ListTables():
        ap.management.TableToDomain(sheet, "Code", "Description", out_gdb, sheet)
        ap.management.Delete(sheet)
        print(f'{sheet} domain created')

    # Assign domains to feature classes
    for fc in ap.ListFeatureClasses():
        for field in ap.ListFields(fc):
            for domain in ap.da.ListDomains():
                if field.name == domain.name:
                    ap.management.AssignDomainToField(fc, field.name, domain.name)
                    print(f'{domain.name} domain assigned to field')

    print('\n>>> DONE >>>')

except Exception as e:
    print(f'An error occurred: {e}')
